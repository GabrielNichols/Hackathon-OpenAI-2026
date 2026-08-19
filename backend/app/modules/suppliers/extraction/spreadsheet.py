from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from .models import (
    ExtractedFieldDTO,
    ExtractionFieldStatus,
    SourceDocumentDTO,
    SupplierExtractionResultDTO,
    SupplierExtractionSchema,
)
from .normalization import (
    AmbiguousValueError,
    normalize_people_quantity,
    normalize_price_to_cents,
)

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_ALIASES = {
    "nome comercial": "trade_name",
    "trade name": "trade_name",
    "categorias": "categories",
    "categoria": "categories",
    "cidades atendidas": "service_cities",
    "cidade atendida": "service_cities",
    "bairros atendidos": "service_districts",
    "bairro atendido": "service_districts",
    "quantidade minima": "minimum_people",
    "minimo de pessoas": "minimum_people",
    "capacidade maxima": "maximum_people",
    "maximo de pessoas": "maximum_people",
    "antecedencia minima horas": "lead_time_hours",
    "lead time horas": "lead_time_hours",
    "emite nota fiscal": "invoice_available",
    "emissao de nf": "invoice_available",
    "forma de precificacao": "pricing_model",
    "modelo de preco": "pricing_model",
    "vegetariano": "vegetarian_supported",
    "vegano": "vegan_supported",
    "sem gluten": "gluten_free_supported",
    "risco de contaminacao cruzada": "cross_contamination_warning",
    "preco base": "base_price_cents",
}
_LIST_FIELDS = {"categories", "service_cities", "service_districts"}
_BOOLEAN_FIELDS = {
    "invoice_available",
    "vegetarian_supported",
    "vegan_supported",
    "gluten_free_supported",
    "cross_contamination_warning",
}


class SpreadsheetExtractionError(ValueError):
    pass


def _fold(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value.casefold())
    without_accents = "".join(
        character for character in decomposed if not unicodedata.combining(character)
    )
    return re.sub(r"[^a-z0-9]+", " ", without_accents).strip()


def _parse_boolean(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    if not isinstance(value, str):
        return None
    folded = _fold(value)
    if folded in {"sim", "yes", "verdadeiro", "disponivel", "suportado"}:
        return True
    if folded in {"nao", "no", "falso", "indisponivel", "nao suportado"}:
        return False
    return None


def _normalize_value(field_name: str, raw_value: Any) -> tuple[Any | None, ExtractionFieldStatus]:
    try:
        if field_name in _LIST_FIELDS:
            if not isinstance(raw_value, str):
                raise TypeError("list field must be text")
            values = [item.strip() for item in re.split(r"[,;\n]", raw_value) if item.strip()]
            if not values:
                raise ValueError("list field is empty")
            return values, ExtractionFieldStatus.EXTRACTED
        if field_name in {"minimum_people", "maximum_people"}:
            return normalize_people_quantity(raw_value), ExtractionFieldStatus.EXTRACTED
        if field_name == "lead_time_hours":
            return normalize_people_quantity(raw_value), ExtractionFieldStatus.EXTRACTED
        if field_name == "base_price_cents":
            return normalize_price_to_cents(raw_value), ExtractionFieldStatus.EXTRACTED
        if field_name in _BOOLEAN_FIELDS:
            normalized = _parse_boolean(raw_value)
            if normalized is None:
                return None, ExtractionFieldStatus.NEEDS_REVIEW
            return normalized, ExtractionFieldStatus.EXTRACTED
        if isinstance(raw_value, str):
            normalized_text = raw_value.strip()
            if not normalized_text:
                raise ValueError("text field is empty")
            return normalized_text, ExtractionFieldStatus.EXTRACTED
        return raw_value, ExtractionFieldStatus.EXTRACTED
    except (AmbiguousValueError, TypeError, ValueError):
        return None, ExtractionFieldStatus.NEEDS_REVIEW


def parse_supplier_spreadsheet(
    *,
    document: SourceDocumentDTO,
    content: bytes,
    schema: SupplierExtractionSchema,
    extracted_at: datetime,
    maximum_cells: int = 100_000,
) -> SupplierExtractionResultDTO:
    """Parse adjacent label/value cells without guessing layout or supplier confirmation."""

    if document.media_type != _XLSX_MEDIA_TYPE:
        raise SpreadsheetExtractionError("deterministic spreadsheet parser requires XLSX")
    if not content:
        raise SpreadsheetExtractionError("spreadsheet content is empty")
    if maximum_cells <= 0:
        raise ValueError("maximum_cells must be positive")

    allowed_fields = {field.name for field in schema.fields}
    matches: dict[str, list[tuple[str, str, Any]]] = defaultdict(list)
    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    visited_cells = 0
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                visited_cells += len(row)
                if visited_cells > maximum_cells:
                    raise SpreadsheetExtractionError("spreadsheet exceeds deterministic cell limit")
                for index, label_cell in enumerate(row[:-1]):
                    if not isinstance(label_cell.value, str):
                        continue
                    field_name = _ALIASES.get(_fold(label_cell.value))
                    if field_name is None or field_name not in allowed_fields:
                        continue
                    value_cell = row[index + 1]
                    if value_cell.value is None:
                        continue
                    matches[field_name].append(
                        (worksheet.title, value_cell.coordinate, value_cell.value)
                    )
    finally:
        workbook.close()

    run_digest = hashlib.sha256(document.document_id.encode() + content).hexdigest()[:16]
    run_id = f"ext_xlsx_{run_digest}"
    fields: list[ExtractedFieldDTO] = []
    for field_definition in schema.fields:
        field_matches = matches.get(field_definition.name, [])
        if not field_matches:
            continue
        source_sheet, source_cell, raw_value = field_matches[0]
        normalized_value, field_status = _normalize_value(field_definition.name, raw_value)
        if len(field_matches) > 1:
            normalized_candidates = [
                _normalize_value(field_definition.name, candidate[2])[0]
                for candidate in field_matches
            ]
            if any(
                candidate != normalized_candidates[0] for candidate in normalized_candidates[1:]
            ):
                raw_value = [candidate[2] for candidate in field_matches]
                normalized_value = None
                field_status = ExtractionFieldStatus.NEEDS_REVIEW
        source_refs = ", ".join(f"{sheet}!{cell}" for sheet, cell, _ in field_matches)
        fields.append(
            ExtractedFieldDTO(
                field_name=field_definition.name,
                value=raw_value,
                normalized_value=normalized_value,
                status=field_status,
                confidence=1.0 if field_status is ExtractionFieldStatus.EXTRACTED else 0.5,
                source_document_id=document.document_id,
                source_page=None,
                source_sheet=source_sheet,
                source_cell_range=source_cell,
                source_excerpt=f"Spreadsheet source: {source_refs}",
                extraction_run_id=run_id,
                confirmed_by=None,
                confirmed_at=None,
                version=1,
            )
        )
    return SupplierExtractionResultDTO(
        extraction_run_id=run_id,
        document_id=document.document_id,
        extracted_at=extracted_at,
        fields=fields,
    )
